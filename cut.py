import io
import sys
from urllib.request import urlopen
sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')


import six
from google.cloud import language
from google.cloud.language import enums
from google.cloud.language import types


from datetime import datetime, timedelta 
import datetime

import os

credential_path = r"Hotel-4f4067fc1229.json"
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = credential_path

from tempfile import NamedTemporaryFile
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook


def syntax_text(text):

    text = text
    client = language.LanguageServiceClient()

    if isinstance(text, six.binary_type):
        text = text.decode('utf-8')
    document = types.Document(
        content=text,
        language = 'zh-Hant',
        type=enums.Document.Type.PLAIN_TEXT)

    tokens = client.analyze_syntax(document).tokens
    final = {}
    for token in tokens:
        part_of_speech_tag = enums.PartOfSpeech.Tag(token.part_of_speech.tag)
       
        if part_of_speech_tag.name in final.keys():
            final[part_of_speech_tag.name].append(token.text.content)
        else:
            final[part_of_speech_tag.name] = [token.text.content]
    
    if not 'VERB' in final.keys():
        final['VERB'] = ''
    if not 'NOUN' in final.keys():
        final['NOUN'] = ''
    if not 'ADJ' in final.keys():
        final['ADJ'] = ''
    return final['VERB'],final['NOUN'],final['ADJ']

def sentiment_text(text):

    # from google.cloud import language
    text = text
    client = language.LanguageServiceClient()
    try:
        text = text.decode('utf-8')
    except AttributeError:
        pass

    document = types.Document(
        content=text,
        language = 'zh-Hant',
        type=enums.Document.Type.PLAIN_TEXT)

    sentiment = client.analyze_sentiment(document).document_sentiment
    return sentiment.score,sentiment.magnitude
    print('Score: {}'.format(sentiment.score))
    print('Magnitude: {}'.format(sentiment.magnitude))

def clean_text(text):

    have_1 = ["零", "壹", "貳", "參", "肆", "伍", "陸", "柒", "捌", "玖"]
    try:
        for word in text[:10]:
            if word in have_1:
                text = text.replace(word, str(have_1.index(word)),1)
    except TypeError:
        return text

    have_2 = ['零','一','二','三','四','五','六','七','八','九']
    for word in text[:10]:
        if word in have_2:
            text = text.replace(word, str(have_2.index(word)),1)
    if '外科' in text:
        text = text.replace('外科', '外客')
    # 把所有國字轉為數字
    text = text.strip()
    print(text)
    # 刪除空白
    find_name(text)

def find_name(text):

    test = ['晚餐', '午餐','早餐', '下午茶']
    word = next((x for x in test if x in text), None)
    if word:
        if word in text[:15]:
            time_def(text,word)

    else:
        test = ['房號', '房客','外客']
        word = next((x for x in test if x in text), None)
        if word:
            if word in text[:15]:
                time_def(text,word)
        else:
                time = ''
                name = ''
                part = ''

                other = text 
                score,magnitude = sentiment_text(text)
                verb, noun, adj = syntax_text(text)
                # wirte_excel(part,time,name,other,verb, noun, adj,score,magnitude)

def time_def(text,cut):
    time = text.split(cut)[0]
    part = ''
    other = cut + text.split(cut)[1] 
    print(time)
    if time == '':
        name,other = cut_name(other) 
        final_time = ''
     
        score,magnitude = sentiment_text(other)
        verb, noun, adj = syntax_text(other)
    
        # return wirte_excel(part,final_time,name,other,verb, noun, adj,score,magnitude)

    for word in time:
        if word.isdigit():
      
            digit = time.index(word)
            
            break
        else:
            digit = None
    if digit:
        part = time[0:digit]
        time = time[digit:]
    print(part)
        
    have = ['月','號','日','早上','早','下午','下','午','晚上','點','分','半','晚','上']

    for word in time:
        if word not in have and word.isdigit() != True:
            time = time.replace(word, '')
            print(word)
            
    print(time)
    if time == '':
        name,other = cut_name(other) 
        final_time = ''
     
        score,magnitude = sentiment_text(other)
        verb, noun, adj = syntax_text(other)
    
        # return wirte_excel(part,final_time,name,other,verb, noun, adj,score,magnitude)

    today = datetime.datetime.now() 
    time = str(today.year)+'年'+time

    if '日' in time:
        time = time.replace('日', '號')
    
    if '點' not in time:
        final_time = cut_time_day(time)
    
    else:
    
        if '月' and '號' in time:
            if '早上' in time:
                time = time.replace('早上', '')
            elif '下午' or '晚上' in time:
                if '下午' in time:
                    new = time[:(time.index('下午')+2)]+str(int(time[(time.index('下午')+2):time.index('點')]) +12)+time[time.index('點'):]
                    time = new.replace('下午', '')
                    print(time)    
                else:
                    new = time[:(time.index('晚上')+2)]+str(int(time[(time.index('晚上')+2):time.index('點')]) +12)+time[time.index('點'):]
                    time = new.replace('晚上', '')
            else:
                time = time
                
        else:
            if '早上' in time:
                temp = list(time)
                station = temp.index('早') 
                time =  "".join(temp[0:(station)])+'號'+"".join(temp[station:]) 
                time = time.replace('早上', '')
            elif '下午' or '晚上' in time:
                if '下午' in time:
                    temp = list(time)
                    station = temp.index('下') 
                    time =  "".join(temp[0:(station)])+'號'+"".join(temp[station:]) 
                    new = time[:(time.index('下午')+2)]+str(int(time[(time.index('下午')+2):time.index('點')]) +12)+time[time.index('點'):]
                    time = new.replace('下午', '')
                    print(time)
                   
                else:
                    temp = list(time)
                    station = temp.index('晚') 
                    time =  "".join(temp[0:(station)])+'號'+"".join(temp[station:]) 
                    new = time[:(time.index('晚上')+2)]+str(int(time[(time.index('晚上')+2):time.index('點')]) +12)+time[time.index('點'):]
                    time = new.replace('晚上', '')
        print(time)
        final_time = cut_time_hour(time)

    name,other = cut_name(other) 
    score,magnitude = sentiment_text(other)
    verb, noun, adj = syntax_text(other)

    # wirte_excel(part,final_time,name,other,verb, noun, adj,score,magnitude)

def cut_time_day(text):
    if '號' in text:
        pass
    else:
        text = text + '號'
    try:
        time = datetime.datetime.strptime(text, '%Y年%m月%d號').date()
    except ValueError:
        return text
    print(time)
    return time

def  cut_time_hour(text):
    if '半' in text:
        try:
            text = text.split('半')[0]
            time = datetime.datetime.strptime(text, '%Y年%m月%d號%H點')
            time +=  datetime.timedelta(minutes=30)
        except ValueError:
            return text
    elif '分' in text:
        try:
            time = datetime.datetime.strptime(text, '%Y年%m月%d號%H點%M分')
        except ValueError:
            return text
    else:
        try:
            test = list(text)
            station = text.index('點')
            if '點' == test[-1]:
                time = datetime.datetime.strptime(text, '%Y年%m月%d號%H點')
            elif test[station+1].isdigit() :
                test.append('分')
                test = "".join(test)
                time = datetime.datetime.strptime(test, '%Y年%m月%d號%H點%M分')
        except ValueError:
            return text

    return time
    print(time)

def cut_name(text):
    test = ['小姐','先生','客人','夫妻']

    word = next((x for x in test if x in text), None)
    if word:
        name = text.split(word,1)[0]+ word
        other = text.split(word,1)[1] # 指切前面的就好 
    else:
        name,other = '', text

    return name,other

# def wirte_excel(part,time,name,other,verb, noun, adj,score,magnitude):

#     wb = Workbook()
    
#     wb = load_workbook('/Users/ginger/Desktop/涵碧樓/sound.xlsx')
#     ws = wb.active
#     # ws['A1'] ='part'
#     # ws['B1'] = 'time'
#     # ws['C1'] = 'name'
#     # ws['D1'] = 'event'
#     # ws['E1'] = 'verb'
#     # ws['F1'] = 'noun'
#     # ws['G1'] = 'adj'
#     # ws['H1'] = 'score'
#     # ws['I1'] = 'magnitude'
#     ws['C1'] ='part'
#     ws['D1'] = 'time'
#     ws['E1'] = 'name'
#     ws['F1'] = 'event'
#     ws['G1'] = 'verb'
#     ws['H1'] = 'noun'
#     ws['I1'] = 'adj'
#     ws['J1'] = 'score'
#     ws['K1'] = 'magnitude'
    
#     if score >0.1:
#         react = '正面'
#     elif score == 0:
#         react = '中立'
#     else:
#         react = '負面'
#     ws.append(['','',part,str(time),name,str(other),str(verb),str(noun),str(adj) ,score,magnitude,react])

#     wb.save('/Users/ginger/Desktop/涵碧樓/sound.xlsx')
#     print('success')



#FIXME:這邊是測試
content = u'東方餐廳2月1日星期六晚餐房客11021 103附加生先生住在13000為那傅先生三位市房價內含的三個晚餐餐點的部分都會喜歡拿，另外有家庭的一瓶甜白酒，客人要求退貨，需要比較冰一點的，那夫人員這邊有幫客人準備一個冰後，並且讓客人先適合，是否是他想要的溫度，然後來復健身有測試了一下，覺得這個溫度很好，那這部分都很喜歡，那份量也很充足。 謝謝我們的服務。'

clean_text(content)
#clean_text('東方餐廳2月1日星期六晚餐房客11021 103附加生先生住在13000為那傅先生三位市房價內含的三個晚餐餐點的部分都會喜歡拿，另外有家庭的一瓶甜白酒，客人要求退貨，需要比較冰一點的，那夫人員這邊有幫客人準備一個冰後，並且讓客人先適合，是否是他想要的溫度，然後來復健身有測試了一下，覺得這個溫度很好，那這部分都很喜歡，那份量也很充足。 謝謝我們的服務。')
# from openpyxl import load_workbook

# wb = load_workbook('/Users/ginger/Desktop/create_sample.xlsx')
# ws = wb.active
# first_column = ws['B']
# for x in range(len(first_column)): 
#     # print(first_column[x].value) 
#     clean_text(first_column[x].value)
